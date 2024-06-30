# rag-stream-manual/backend/utils/vectorizer.py

import os
import requests
import pandas as pd
import numpy as np
from io import BytesIO
from pypdf import PdfReader
from docx import Document
from openpyxl import load_workbook
from langchain.text_splitter import CharacterTextSplitter
from langchain_openai import OpenAIEmbeddings
from dotenv import load_dotenv
import logging

logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

load_dotenv()

OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
CHUNK_SIZE = int(os.getenv("CHUNK_SIZE", 15))
CHUNK_OVERLAP = int(os.getenv("CHUNK_OVERLAP", 0))
SEPARATOR = os.getenv("SEPARATOR", "\n\n")
XLSX_SEPARATOR = os.getenv("XLSX_SEPARATOR", "\n")
CSV_OUTPUT_DIR = os.getenv("CSV_OUTPUT_DIR", "../data/csv/all")

is_docker = os.getenv("IS_DOCKER", "false").lower() == "true"
S3_DB_URL = os.getenv("S3_DB_INTERNAL_URL" if is_docker else "S3_DB_EXTERNAL_URL", "http://localhost:9001")

logger.info(f"Using S3_DB_URL: {S3_DB_URL}")

embeddings = OpenAIEmbeddings(model="text-embedding-3-large")

def normalize_vector(vector):
    return vector / np.linalg.norm(vector)

def get_files_from_s3(file_type):
    try:
        response = requests.get(f"{S3_DB_URL}/data/{file_type}", timeout=10)
        response.raise_for_status()
        files = response.json()
        valid_files = [file for file in files if file.endswith(f'.{file_type}')]
        logger.info(f"Found {len(valid_files)} {file_type.upper()} files: {valid_files}")
        return valid_files
    except Exception as e:
        logger.error(f"Error fetching {file_type.upper()} files: {str(e)}")
        return []

def fetch_file_content(file_type, file_name):
    try:
        response = requests.get(f"{S3_DB_URL}/data/{file_type}/{file_name}", stream=True, timeout=10)
        response.raise_for_status()
        logger.info(f"Successfully fetched {file_type.upper()} file: {file_name}")
        return response.content
    except Exception as e:
        logger.error(f"Error fetching {file_type.upper()} file {file_name}: {str(e)}")
        return None

def extract_text(content, file_type):
    if file_type == 'pdf':
        pdf = PdfReader(BytesIO(content))
        return [(str(i+1), page.extract_text()) for i, page in enumerate(pdf.pages)]
    elif file_type == 'xlsx':
        workbook = load_workbook(filename=BytesIO(content), read_only=True)
        return [(sheet.title, XLSX_SEPARATOR.join([" ".join([str(cell.value) for cell in row if cell.value is not None]) for row in sheet.iter_rows()])) for sheet in workbook]
    elif file_type == 'docx':
        return [('1', SEPARATOR.join([para.text for para in Document(BytesIO(content)).paragraphs]))]

def process_file(file_type, file_name):
    content = fetch_file_content(file_type, file_name)
    if content is None:
        return pd.DataFrame()

    texts = extract_text(content, file_type)
    text_splitter = CharacterTextSplitter(
        chunk_size=CHUNK_SIZE,
        chunk_overlap=CHUNK_OVERLAP,
        separator=XLSX_SEPARATOR if file_type == 'xlsx' else SEPARATOR
    )

    processed_data = []
    for location, text in texts:
        for chunk in text_splitter.split_text(text):
            vector = embeddings.embed_query(chunk)
            processed_data.append({
                'file_name': file_name,
                'file_type': file_type.upper(),
                'location': location,
                'manual': chunk,
                'manual_vector': normalize_vector(vector).tolist()
            })

    return pd.DataFrame(processed_data)

def main():
    all_processed_data = pd.concat([process_file(file_type, file_name)
                                    for file_type in ['pdf', 'xlsx', 'docx']
                                    for file_name in get_files_from_s3(file_type)],
                                    ignore_index=True)

    if not all_processed_data.empty:
        os.makedirs(CSV_OUTPUT_DIR, exist_ok=True)
        output_file = os.path.join(CSV_OUTPUT_DIR, "all_documents_vector_normalized.csv")
        all_processed_data.to_csv(output_file, index=False)
        logger.info(f"All processed data saved to {output_file}")
    else:
        logger.warning("No data processed.")

if __name__ == "__main__":
    main()
